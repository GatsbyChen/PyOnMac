from openpyxl import load_workbook
import pandas as pd

df = pd.DataFrame()
wb = load_workbook(r"C:\Users\2067\PycharmProjects\test1\Dalian_Enterprise\InfoSheet.xlsx")
sh = wb["Sheet1"]

print(sh["C30"].value)
# print(sh.cell("T30"))

positionlist = []
with open("ExtractionPosition.txt","r",encoding="utf8") as f:
    lines = f.readlines()
    for i in lines:
        positionlist.append(i.strip())
print(positionlist)
positionlist = ['C30', 'T30', 'C31', 'T31', 'C32', 'I32', 'P32', 'V32', 'C33', 'T33', 'C34', 'M34', 'W34', 'C35', 'T35', 'C37', 'W37', 'C38', 'W38', 'C39', 'S39', 'C40', 'K40', 'S40', 'C41', 'J41', 'Q41', 'V41', 'C42', 'T42', 'C43', 'A48', 'F48', 'I48', 'L48', 'O48', 'S48', 'V48', 'A49', 'F49', 'I49', 'L49', 'O49', 'S49', 'V49', 'C50', 'T50', '']

infolist = []
for i in positionlist:
    infolist.append(sh[""+i+""].value)
print(infolist)